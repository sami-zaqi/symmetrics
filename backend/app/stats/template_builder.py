import io

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Font

from app.core.schemas import DataSchema


def build_template_xlsx(schema: DataSchema) -> bytes:
    """Two-sheet data-collection template: 'Data' has the correct column
    headers ready to fill in, 'Panduan Kode' is the human-readable answer
    key (category codes, questionnaire item groupings, missing-value symbol)
    so the student fills the sheet consistently from the very first row."""
    wb = Workbook()
    data_sheet = wb.active
    data_sheet.title = "Data"

    bold = Font(bold=True)
    for col_idx, var in enumerate(schema.variables, start=1):
        cell = data_sheet.cell(row=1, column=col_idx, value=var.name)
        cell.font = bold
        cell.comment = Comment(f"{var.label} ({var.scale})", "Symmetrics")
        data_sheet.column_dimensions[cell.column_letter].width = max(12, len(var.name) + 4)

    guide = wb.create_sheet("Panduan Kode")
    row = 1
    guide.cell(row=row, column=1, value="Panduan Pengisian Data").font = Font(bold=True, size=13)
    row += 2

    if schema.missing_value_symbol:
        guide.cell(
            row=row, column=1,
            value=f'Data kosong/hilang: tulis "{schema.missing_value_symbol}" pada sel (jangan dikosongkan begitu saja).',
        )
    else:
        guide.cell(row=row, column=1, value="Data kosong/hilang: cukup biarkan sel kosong.")
    row += 2

    categorical_vars = [v for v in schema.variables if v.categories]
    if categorical_vars:
        guide.cell(row=row, column=1, value="Kode Kategori").font = Font(bold=True)
        row += 1
        for var in categorical_vars:
            guide.cell(row=row, column=1, value=f"{var.name} ({var.label}):").font = Font(bold=True)
            row += 1
            for cat in var.categories or []:
                guide.cell(row=row, column=1, value=f"  {cat.value} = {cat.label}")
                row += 1
            row += 1

    if schema.constructs:
        guide.cell(row=row, column=1, value="Kelompok Item Kuesioner (Konstruk)").font = Font(bold=True)
        row += 1
        for c in schema.constructs:
            guide.cell(row=row, column=1, value=f"{c.name}: {', '.join(c.items)}")
            row += 1

    guide.column_dimensions["A"].width = 65

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
